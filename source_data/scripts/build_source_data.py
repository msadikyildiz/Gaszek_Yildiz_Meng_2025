#!/usr/bin/env python3
"""Assemble source_data.xlsx: one sheet per display item. Populates the new revision
figures whose underlying values live in verified analysis CSVs; stubs parquet-derived and
contributor figures with a clear source/owner note so they can be filled at figure-lock.

NOTE: the final workbook is release-gated on one still-pending contributor sheet
(Fig 6 - F.M./Sophia/Alberto, which needs the PF13354 MSA + py-mfdca to run). S18 is now
reproduced from Meng's pipeline (src/graph_analysis/s18_peak_robustness/reproduce_s18.py).
Running this script before Fig 6 lands produces a workbook with that sheet (plus Fig1F,
which has no plotted data) stubbed out -- see source_data/README.md.
"""
import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = Path(__file__).resolve().parent


def _repo_root(_p):
    """Walk up from this file to the repo root (dir with data/processed/Epistasis_Combined.parquet)."""
    for _a in (_p, *_p.parents):
        if (_a / "data" / "processed" / "Epistasis_Combined.parquet").exists():
            return _a
    raise FileNotFoundError("repo root not found from " + str(_p))


REPO = _repo_root(Path(__file__).resolve())
OUTDIR = HERE.parent
OUTDIR.mkdir(exist_ok=True)
OUT = OUTDIR / "source_data.xlsx"

# figure -> list of (section label, csv relpath under the repo root)
PD = "source_data/derived"
DATA = {
 "Fig1E global fitness distributions": [
    ("Per-concentration distribution summary + WT/dead reference lines", f"{PD}/fig1E_distribution_summary.csv"),
    ("Histogram counts per drug x concentration (reproduces the ridgelines)", f"{PD}/fig1E_histogram_counts.csv")],
 "Fig2 landscape graphs": [
    ("Coarse-grained landscape-graph NODES per plotted concentration (fitness = median at that conc; n_genotypes = merged group size; is_peak = out-degree 0; contains_wildtype)", f"{PD}/fig2_nodes.csv"),
    ("Landscape-graph EDGES per plotted concentration: source -> target oriented by the sign of the net aggregate transition flow (forward minus reverse) after neutral-merging supernodes - predominantly toward higher fitness, though a few edges (3 of 2873) point to a lower-median-fitness supernode where aggregation reverses the net flow; weight = |net flow| = |summed exp(|Δfitness|) over forward minus reverse single-mutation transitions| (net magnitudes, not raw exp(|Δfitness|) - near-balanced pairs are small, min ~0.09); count = number of merged transitions", f"{PD}/fig2_edges.csv")],
 "Fig3 landscape signatures": [
    ("A: overlay points (WT, dead, 7 follow-ups, clinical isolates) - fitness at AMP 195/781 & AZT 36/108/324", f"{PD}/fig3A_overlay_points.csv"),
    ("B: top-1% sequence-logo residue frequencies at AMP 781 (letter height + colour)", f"{PD}/fig3B_top1pct_logo_freq_amp.csv"),
    ("B: top-1% sequence-logo residue frequencies at AZT 36", f"{PD}/fig3B_top1pct_logo_freq_azt.csv"),
    ("C/D: median fitness vs number of mutations (the black line), AMP 781 / AZT 36", f"{PD}/fig3CD_median_by_nmut.csv")],
 "Fig4A-B single-mutant fitness": [
    ("Single-mutant AUC-fitness, n=3 selection replicates + mean + s.d. (AMP 781 / AZT 36)", f"{PD}/fig4AB_single_mutant_replicates.csv")],
 "Fig4C-D double-mutant fitness": [
    ("AUC-fitness for every single/double combination of the 18 mutations (long form; none = WT)", f"{PD}/fig4CD_double_mutant_fitness.csv")],
 "Fig4E-F measured vs predicted": [
    ("Per-order fit summary: Pearson r, R^2, RMSD, slope, intercept (AMP 781 / AZT 36)", f"{PD}/fig4EF_summary.csv")],
 "Fig5 SHAP interpretation": [
    ("C: mean|SHAP| and mean SHAP per mutation (AMP / AZT)", f"{PD}/fig5C_mean_abs_shap.csv"),
    ("Seed-robustness: per-mutation mean|SHAP| across 4 split seeds", f"{PD}/fig5_shap_seed_stability.csv")],
 "S12 R2 by order across conc": [
    ("Per drug x concentration x order (<=1..6): Pearson R^2 + RMSD", f"{PD}/figS12_r2_by_order_summary.csv")],
 "S7-S8 IC50 validation": [
    ("S7: monoculture log10(IC50) vs mean AUC-fitness per variant (AMP 20260407 / AZT 20260307)", f"{PD}/figS7_ic50_vs_meanfitness.csv"),
    ("S7: Pearson r/p per drug (DD plotted but excluded - mean_fitness null)", f"{PD}/figS7_pearson_correlation.csv"),
    ("S8: per-concentration AUC-fitness vs log10(IC50), long form (0-drug excluded)", f"{PD}/figS8_ic50_vs_fitness_per_conc.csv"),
    ("S8: per-concentration Pearson r/p", f"{PD}/figS8_pearson_by_conc.csv")],
 "S9-S10 order decomposition": [
    ("Per-order R^2 / cumulative R^2 / dR^2 / RMSD (all conc)", "analysis/s09_s10_epistatic_order/data/order_decomposition.csv"),
    ("Matched-stringency tiers (order R^2/RMSD)", "analysis/s09_s10_epistatic_order/data/matched_stringency_summary.csv")],
 "S11 pairwise epistasis": [
    ("Biochemical pairwise epistasis per mutation pair x concentration", "analysis/s11_s12_concentration_grid/data/pairwise_mean_fitness.csv")],
 "S13 learning curves": [
    ("RMSD vs training fraction per model x feature-set (mean+/-std over seeds)", "analysis/s13_model_comparison/data/learning_curves_summary.csv")],
 "S14 LOMO + Hamming holdout": [
    ("Leave-one-mutation-out holdout (RMSD/R2 vs random)", "analysis/s14_mutation_holdout/data/lomo_results.csv"),
    ("Hamming-distance-stratified holdout", "analysis/s14_mutation_holdout/data/hamming_results.csv"),
    ("Classification metrics (block holdout)", "analysis/s14_mutation_holdout/data/classification_results.csv")],
 "S15 classification metrics": [
    ("Sensitivity / positive-rate at fixed thresholds", "analysis/s15_classification_metrics/results_table.csv")],
 "S16 replicate CV": [
    ("Per-concentration replicate %CV (viable subset)", "analysis/s16_replicate_reproducibility/data/summary_by_conc_viable.csv"),
    ("Per-concentration replicate %CV (all genotypes / pooled)", "analysis/s16_replicate_reproducibility/data/summary_by_conc.csv"),
    ("Replicate-pair Pearson r (per drug x concentration)", "analysis/s16_replicate_reproducibility/data/replicate_pair_correlations.csv")],
 "S17 dose-response low-count": [
    ("Per-variant dose-response AUC, replicate reads, low-count flags", "analysis/s17_dose_response_low_count/data/per_variant_fitness.csv"),
    ("Highlighted-variant categories", "analysis/s17_dose_response_low_count/data/highlighted_variants.csv")],
 "Fig4G-H R2 by order": [
    ("Cumulative R^2 for orders 1-6 per drug x concentration", "analysis/s11_s12_concentration_grid/results_table.csv")],
 "S18 peak-advantage + threshold robustness": [
    ("C: neutral-threshold robustness — has_global_peak per AZT concentration x neutral threshold (0.15-0.45)", "src/graph_analysis/s18_peak_robustness/source_data/figS18C_neutral_threshold_matrix.csv"),
    ("A: AZT-12 global-peak-supernode fitness advantage over external neighbours, box stats per concentration x neighbour distance", "src/graph_analysis/s18_peak_robustness/source_data/figS18A_azt12_peak_advantage_boxstats.csv"),
    ("B: AZT-108 global-peak-supernode fitness advantage, box stats per concentration x neighbour distance", "src/graph_analysis/s18_peak_robustness/source_data/figS18B_azt108_peak_advantage_boxstats.csv")],
}

# figure -> (source note, owner) for not-yet-filled sheets
STUB = {
 "Fig6 - TODO Morcos lab":   ("DCA code folded in at src/evolutionary_statistics/ (panels A/B/C/G/H: family logo, effective alphabet, Hamiltonian C/G/H); LGL D/E/F via github.com/morcoslab/LGL-VAE. Numbers pending: add MSAs/PF13354_ga.fasta + install py-mfdca to run + extract, and obtain LGL-VAE outputs — F.M./Sophia/Alberto", "F.M./Sophia/Alberto"),
 "Fig1F no plotted data":    ("Structural render + 2D chemical structures — no plotted data; ChemDraw supplied as artwork", "IG/DS"),
}

# figure -> provenance / raw-data-location note (written at top of the sheet)
NOTES = {
 "Fig2 landscape graphs": (
    "Reproduced deterministically from D. Meng's fitness-landscape-graph builder "
    "(neutral_threshold=0.4, tiny_initial_threshold=0.02, large_edge_threshold=5.5, "
    "num_forbidden_pairs=1) run on the design-filtered per-genotype AUC data. Nodes are "
    "neutral-merged supernodes (per-concentration median fitness); edges are net single-mutation "
    "transition flows, oriented by the sign of forward-minus-reverse aggregate weight "
    "(predominantly, but not strictly, toward higher fitness). Plotted concentrations: AMP "
    "12.2/48.8/195 and AZT 12/36/108 µg/mL; all 55,296 design genotypes are partitioned across "
    "the nodes at each concentration."),
 "Fig3 landscape signatures": (
    "Panel A is a 2x3 hex-bin density of AZT (36/108/324) vs AMP (195/781) fitness for all 55,296 "
    "genotypes; bulk per-genotype values = companion derived/fig3A_all_genotypes.csv (also in "
    "Epistasis_Combined.parquet), overlay points tabulated here. Panel B letter height = residue "
    "frequency among the top 1% (n=553); logo_color (black=WT, red=enriched above the threshold, "
    "grey=other) reproduces the panel colouring (legend groups R164H with R164S/N under AZT although "
    "R164H falls just below threshold). Panels C/D per-genotype (n_mut, fitness) = companion "
    "fig3CD_all_genotypes.csv; the plotted median line is tabulated here."),
 "Fig4A-B single-mutant fitness": (
    "AUC-fitness of each single mutant with its three independent selection replicates (n=3), mean and "
    "s.d., at the reference dose; WT included (dashed reference line). Replicates from *_auc_long_df.parquet."),
 "Fig4C-D double-mutant fitness": (
    "Heat-map values: AUC-Fitness (the Epistasis_Combined per-genotype value = median of the 3 replicates) "
    "for every single (mutation_2 = none) and double combination of the 18 mutations at the reference dose. "
    "This is fitness, not epistasis (the pairwise-epistasis heatmap is a separate, non-manuscript panel)."),
 "Fig5 SHAP interpretation": (
    "Deterministic re-computation of the published Fig 5 SHAP pipeline (the figure itself is unchanged): "
    "Epistasis_Combined at the reference dose -> 18 one-hot mutation features -> LGBMRegressor(n_estimators="
    "100, learning_rate=0.1, seed 42) trained on a 10% split -> shap.TreeExplainer over the 90% test set. "
    "The original notebooks used an unseeded split; the attributions do not depend on it - per-mutation "
    "mean|SHAP| Spearman rho vs seed 42 = 0.95-0.99 across four seeds, max drift <=0.018 (seed-stability "
    "block below). The 5A/5B heatmap matrices (top-553 test genotypes x 18 mutations, ordered most->least "
    "resistant) and the full test-set SHAP behind the beeswarm are the companion files "
    "fig5A/fig5B_shap_top553_*.csv and fig5_shap_all_test_*.csv."),
 "S7-S8 IC50 validation": (
    "Monoculture IC50 (four-parameter Hill fit) vs pooled AUC-fitness for the validation variants. S7 "
    "correlates log10(IC50) against mean AUC-fitness (AMP r=0.88, n=13; AZT r=0.80, n=18; the dead "
    "double-mutant DD is plotted but has null fitness and is excluded from r). S8 correlates IC50 against "
    "AUC-fitness at each non-zero concentration. Source: validation/src/processed/"
    "{20260407 AMP, 20260307 AZT}/xref_expanded_df.parquet."),
 "Fig1E global fitness distributions": (
    "Ridgeline of AUC-fitness per drug x concentration (KDE in the panel; the histogram counts here "
    "reproduce each ridge). Black dashed = TEM-1(WT) median, blue dashed = TEM-1(dead) median. Full "
    "per-genotype AUC-fitness values: Epistasis_Combined.parquet (Zenodo 10.5281/zenodo.21442350)."),
 "Fig4E-F measured vs predicted": (
    "R^2 = squared Pearson correlation of measured vs cumulative-order-K predicted fitness at the "
    "reference dose. Per-genotype (measured, predicted<=1..6) pairs for all 55,296 genotypes are the "
    "companion file derived/fig4EF_measured_vs_predicted.csv and are in Epistasis_Combined.parquet."),
 "S12 R2 by order across conc": (
    "Squared-Pearson R^2 and RMSD of measured vs cumulative-order-K predicted fitness across the full "
    "concentration grid (0-drug panels excluded). Reproduces analysis/s11_s12_concentration_grid exactly. "
    "Per-genotype pairs are in Epistasis_Combined.parquet."),
 "S18 peak-advantage + threshold robustness": (
    "Reproduced from D. Meng's fitness-landscape-graph pipeline (src/graph_analysis/s18_peak_robustness/, "
    "driver reproduce_s18.py) on this repo's AUC data. C: has_global_peak (min_group_size=12) across the "
    "AZT neutral-threshold sweep 0.15-0.45 x 8 concentrations. A/B: fitness advantage (group member minus "
    "external neighbour, up to 2 mutations) of the rank-0 global-peak supernode at AZT 12 / AZT 108 "
    "(neutral_threshold=0.40), summarised as box statistics per concentration x neighbour distance; the "
    "companion figS18{A,B}_*_peak_group_genotypes.csv list the peak-group genotypes behind the panel logos."),
}

HEAD=Font(bold=True); TITLE=Font(bold=True, size=12)
NOTEFILL=PatternFill("solid", fgColor="FFF3CD"); HDRFILL=PatternFill("solid", fgColor="D9E1F2")

wb=Workbook(); wb.remove(wb.active)

# index sheet
idx=wb.create_sheet("Index")
idx["A1"]="Source Data — index"; idx["A1"].font=TITLE
idx.append([]); idx.append(["Figure sheet","Status","Primary source"])
for c in idx[3]: c.font=HEAD; c.fill=HDRFILL
for name in DATA:
    sources = [rel for _label, rel in DATA[name]]
    present = [rel for rel in sources if (REPO / rel).exists()]
    if len(present) == len(sources):
        status = "populated"
    elif not present:
        status = "missing (run builders)"
    else:
        status = f"partial ({len(present)}/{len(sources)})"
    idx.append([name, status, sources[0]])
for name,(note,owner) in STUB.items():
    status = "N/A" if "no plotted data" in note else f"TODO ({owner})"
    idx.append([name, status, note])
idx.column_dimensions["A"].width=30; idx.column_dimensions["B"].width=16; idx.column_dimensions["C"].width=90

def add_csv(ws, row, label, relpath):
    ws.cell(row,1,label).font=HEAD
    p=REPO/relpath
    ws.cell(row+1,1,f"source: {relpath}").font=Font(italic=True, size=9)
    if not p.exists():
        ws.cell(row+2,1,"(file missing)"); return row+4
    with open(p) as fh: rows=list(csv.reader(fh))
    hr=row+2
    for j,v in enumerate(rows[0],1):
        c=ws.cell(hr,j,v); c.font=HEAD; c.fill=HDRFILL
    for i,r in enumerate(rows[1:],1):
        for j,v in enumerate(r,1): ws.cell(hr+i,j,v)
    return hr+len(rows)  # next free row after a gap

for name, sections in DATA.items():
    ws=wb.create_sheet(re.sub(r'[\\/?*\[\]:]', '-', name)[:31])
    ws.cell(1,1,name).font=TITLE
    r=3
    note=NOTES.get(name)
    if note:
        c=ws.cell(2,1,note); c.font=Font(italic=True, size=9); c.alignment=Alignment(wrap_text=True)
        ws.column_dimensions["A"].width=110
        r=4
    for label, rel in sections:
        r=add_csv(ws, r, label, rel)+1

for name,(note,owner) in STUB.items():
    ws=wb.create_sheet(re.sub(r'[\\/?*\[\]:]', '-', name)[:31])
    ws.cell(1,1,name.split(" - ")[0].replace(" N/A","")).font=TITLE
    body = f"N/A — no source data. {note}" if "no plotted data" in note else f"TO FILL ({owner}): {note}"
    c=ws.cell(3,1,body); c.fill=NOTEFILL; c.alignment=Alignment(wrap_text=True)
    ws.column_dimensions["A"].width=110

wb.save(OUT)
print(f"wrote {OUT}")
_fully = [n for n, secs in DATA.items() if all((REPO / rel).exists() for _l, rel in secs)]
print(f"sheets: {len(wb.sheetnames)}  ({len(_fully)}/{len(DATA)} data sheets fully populated, "
      f"{len(DATA) - len(_fully)} missing/partial + {len(STUB)} stub + Index)")
print("data sheets:", ", ".join(DATA.keys()))
